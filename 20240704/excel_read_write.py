import openpyxl  # openpyxl引入模块


# 读取 excel文件
def read_to_excel(path: str, sheet: str, info):
    # 实例化一个workbook对象
    workbook = openpyxl.load_workbook(path)
    # 获取excel文件内的那一个sheet
    data = workbook[sheet]
    # 定义个要输出的总数组
    return_data = []

    # data.rows 为表格内的每一行数据
    # 循环获取表格内的每一行数据
    for index, row in enumerate(data.rows):
        # 判断index为0那一行跳出本次循环忽略本次数据 就不会再添加了
        if index == 0:
            continue

        # 定义一个空的数组用来存放每一行数据单元格的数据
        return_row = []

        for col_index, col_value in enumerate(row):
            # 获取单元格数据 追加到return_row
            # 定义一个{}
            col_obj = {}
            # 通过col_index索引 使info中key的名字和col_value.value能匹配
            # info[col_index]获取key
            # 把 col_value.value赋值给col_obj[info[col_index]]
            col_obj[info[col_index]] = col_value.value
            return_row.append(col_obj)

        return_data.append(return_row)
    # 把遍历出来得每一行数据数据return_row 追加到总数组 return_data中 然后输出

    return return_data


def write_to_excel(path: str, sheetStr, info, data):
    #     实例化一个workbook对象
    workbook = openpyxl.load_workbook(path)
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet设置一个title
    sheet.title = sheetStr

    # 添加表头（不需要表头可以不用加）
    data.insert(0, list(info))
    # 开始遍历数组
    for row_index, row_item in enumerate(data):

        for col_index, col_item in enumerate(row_item):
            # 如果有表头 第二种数据格式就要价格判断单独处理表头
            if row_index == 0:
                # 直接写入 col_item的值
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
            else:
                # 获取字典中 {'name': 'John Brown'}, {'age': 18}, {'address': 'New York No. 1 Lake Park'}
                # list数组列表后为 ['John Brown',18,'New York No. 1 Lake Park']
                arr2 = list(col_item.values())
                # 数组转换为字符串
                str2 = ','.join(str(i) for i in arr2)
                # 写入
                sheet.cell(row=row_index + 1, column=col_index + 1, value=str2)

    # 写入excel文件 如果path路径的文件不存在那么就会自动创建
    workbook.save(path)
    print('写入成功')


if __name__ == '__main__':
    # path 访问文件的路径
    path = r'D:\2.xlsx'
    # Excel 中sheet 的名字
    sheet = 'Sheet1'
    # 在这里定义一个包含key的数组
    info = ['name', 'age', 'address']
    print(read_to_excel(path, sheet, info))

    # 数据结构1Excel 中sheet 的名字
    sheetStr = 'Sheet2'

    info1 = ['name', 'age', 'address']
    # 数据结构2数据
    writeData = read_to_excel(path, sheet, info)
    # 执行
    write_to_excel(path, sheetStr, info1, writeData)
