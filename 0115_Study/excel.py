import openpyxl

wb = openpyxl.load_workbook("data.xlsx")
print(type(wb))
print(wb.sheetnames)

sheet = wb["Sheet1"]
print(sheet)
print(type(sheet))
print(sheet.title)
print("\n")
print(sheet["A2"])
print(sheet["A2"].value)
b = sheet["b2"]
print("b的列名： "+str(b.column))
print(b.coordinate+" is "+str(b.value))

print(sheet.cell(row=2,column=2))
print(sheet.cell(row=2,column=2).value)

for i in range(1,8,2):
    print(i,sheet.cell(row=i,column=2).value)

sheet.max_row
print(sheet.max_row)
print(sheet.max_column)

# tuple(sheet["A1":"D4"])

# for cellij in sheet["A1":"E4"]:
#     for cellj in cellij:
#         print(cellj.coordinate,cellj.value)
#     print("---行结束---")

print(sheet.columns)
#
# for cellij in sheet.columns[2]:
#     print(cellij.value)


