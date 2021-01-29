import openpyxl

#打开工作簿
wb = openpyxl.load_workbook("updateExcel.xlsx")
#打开工作表1
sheet = wb["Sheet1"]
#价格更新字典
PRICE_UPDATES={"苹果":"1.19","橘子":"3.07","香蕉":"1.27"}
print(PRICE_UPDATES["苹果"])
print(PRICE_UPDATES["橘子"])
print(PRICE_UPDATES["香蕉"])

#遍历所有行，并更新价格
for row in range(2,sheet.max_row):
    produceName = sheet.cell(row=row,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=row,column=2).value = PRICE_UPDATES[produceName]

wb.save('updateExcel.xlsx')

#打开工作簿
wb = openpyxl.load_workbook("updateExcel.xlsx")
#打开工作表1
sheet = wb.active
sheet.merge_cells("A8:D12")
sheet["A8"] = "这是一个合并单元格"
sheet.row_dimensions[1].height = 70
sheet["A7"] =700
sheet["B7"] =800
sheet["C7"] = "=SUM(A7:B7)"
wb.save('updateExcel.xlsx')

