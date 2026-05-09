"""
电话费汇总程序
功能：汇总所有月份的电话费Excel表格，根据姓名汇总每月的付费总额
输出：包含每月明细和总计的汇总Excel文件
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from collections import defaultdict


def extract_month_from_filename(filename):
    """从文件名中提取月份信息"""
    # 匹配格式：电话费2025年5月.xlsx 或 电话费2026年01月.xlsx
    match = re.search(r'(\d{4})年(\d{1,2})月', filename)
    if match:
        year = match.group(1)
        month = match.group(2).zfill(2)  # 补齐为2位数
        return f"{year}年{month}月"
    return None


def read_excel_file(filepath):
    """读取单个Excel文件，返回姓名和付费总额的字典"""
    data = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)  # 添加data_only=True以读取公式计算结果
        ws = wb.active
        
        # 遍历所有行（跳过标题行）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                name = row[2]  # 姓名列
                amount = row[3]  # 付费总额列
                
                if name and amount:
                    try:
                        data[name] = float(amount)
                    except (ValueError, TypeError):
                        # 如果转换失败，跳过该行
                        pass
        
        wb.close()
    except Exception as e:
        print(f"读取文件 {filepath} 时出错: {e}")
    
    return data


def collect_all_data(folder_path):
    """收集所有Excel文件的数据"""
    # 数据结构：{姓名: {月份: 金额}}
    all_data = defaultdict(dict)
    months_list = []
    
    # 获取所有xlsx文件
    excel_files = [f for f in os.listdir(folder_path) 
                   if f.endswith('.xlsx') and not f.startswith('~$')]
    
    # 按月份排序
    excel_files.sort(key=lambda x: extract_month_from_filename(x) or '')
    
    print(f"找到 {len(excel_files)} 个Excel文件:")
    for filename in excel_files:
        month = extract_month_from_filename(filename)
        if month:
            print(f"  - {filename} ({month})")
            filepath = os.path.join(folder_path, filename)
            month_data = read_excel_file(filepath)
            
            if month not in months_list:
                months_list.append(month)
            
            # 合并数据
            for name, amount in month_data.items():
                all_data[name][month] = amount
    
    return all_data, months_list


def create_summary_excel(all_data, months_list, output_path):
    """创建汇总Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "电话费汇总"
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 构建表头
    headers = ["姓名"] + months_list + ["总计"]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # 填充数据
    total_by_month = defaultdict(float)  # 每月总计
    grand_total = 0  # 总合计
    
    for name in sorted(all_data.keys()):
        row_data = [name]
        name_total = 0
        
        for month in months_list:
            amount = all_data[name].get(month, 0)
            row_data.append(amount)
            total_by_month[month] += amount
            name_total += amount
        
        row_data.append(name_total)  # 个人总计
        grand_total += name_total
        
        ws.append(row_data)
        
        # 设置数据行样式
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # 添加总计行
    total_row = ["每月合计"]
    for month in months_list:
        total_row.append(total_by_month[month])
    total_row.append(grand_total)
    
    ws.append(total_row)
    
    # 设置总计行样式
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FF0000")
        cell.fill = openpyxl.styles.PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    # 保存文件
    wb.save(output_path)
    wb.close()
    
    return len(all_data), grand_total


def main():
    """主函数"""
    folder_path = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(folder_path, "电话费汇总结果.xlsx")
    
    print("=" * 60)
    print("电话费汇总程序")
    print("=" * 60)
    print(f"\n工作目录: {folder_path}\n")
    
    # 收集所有数据
    print("正在读取Excel文件...")
    all_data, months_list = collect_all_data(folder_path)
    
    if not all_data:
        print("\n未找到任何数据！")
        return
    
    print(f"\n共找到 {len(all_data)} 个人的数据")
    print(f"时间范围: {months_list[0]} 至 {months_list[-1]}")
    
    # 创建汇总文件
    print("\n正在生成汇总Excel文件...")
    person_count, grand_total = create_summary_excel(all_data, months_list, output_file)
    
    print("\n" + "=" * 60)
    print("汇总完成！")
    print("=" * 60)
    print(f"输出文件: {output_file}")
    print(f"总人数: {person_count} 人")
    print(f"总金额: {grand_total:.2f} 元")
    print("=" * 60)


if __name__ == "__main__":
    main()
