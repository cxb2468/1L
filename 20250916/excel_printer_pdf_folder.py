import sys
import os
import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QListWidget, QComboBox, QLabel, QFileDialog, QLineEdit,
    QTextEdit, QCheckBox, QMessageBox, QGroupBox, QProgressBar
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from openpyxl import load_workbook
import win32print
import subprocess

class PrinterWorker(QThread):
    # 定义信号，用于线程与主线程通信
    print_status = pyqtSignal(str)  # 打印状态信号
    log_message = pyqtSignal(str)   # 日志信息信号
    progress_update = pyqtSignal(int)  # 进度更新信号
    
    def __init__(self, file_path, sheet_names, printer_name, output_dir=None):
        super().__init__()
        self.file_path = file_path
        self.sheet_names = sheet_names
        self.printer_name = printer_name
        self.output_dir = output_dir
        
    def run(self):
        try:
            self.log_message.emit(f"开始处理任务: {self.file_path}")
            
            # 如果没有指定工作表，则处理所有工作表
            if not self.sheet_names:
                workbook = load_workbook(self.file_path)
                self.sheet_names = workbook.sheetnames
                workbook.close()
            
            self.log_message.emit(f"将要处理的工作表: {', '.join(self.sheet_names)}")
            
            # 处理每个工作表
            total_sheets = len(self.sheet_names)
            generated_files = []
            
            for i, sheet_name in enumerate(self.sheet_names):
                progress = int((i / total_sheets) * 100)
                self.progress_update.emit(progress)
                
                self.print_status.emit("正在处理")
                self.log_message.emit(f"正在处理工作表: {sheet_name}")
                
                try:
                    if "Adobe PDF" in self.printer_name or "Microsoft Print to PDF" in self.printer_name:
                        # 处理PDF打印机 - 直接导出为PDF
                        pdf_file = self.export_to_pdf(sheet_name)
                        generated_files.append(pdf_file)
                        self.log_message.emit(f"已生成PDF文件: {pdf_file}")
                    else:
                        # 处理普通打印机 - 发送打印任务
                        self.print_to_physical_printer(sheet_name)
                    
                    self.print_status.emit("处理成功")
                    self.log_message.emit(f"工作表 {sheet_name} 处理完成")
                except Exception as e:
                    self.print_status.emit("处理失败")
                    self.log_message.emit(f"工作表 {sheet_name} 处理失败: {str(e)}")
                    
                # 稍微延迟一下，避免过快操作
                self.msleep(500)
                
            self.progress_update.emit(100)
            self.log_message.emit(f"所有工作表处理完成，共生成 {len(generated_files)} 个PDF文件")
            if generated_files:
                self.log_message.emit("生成的文件列表:")
                for file in generated_files:
                    self.log_message.emit(f"  - {file}")
                    
        except Exception as e:
            self.print_status.emit("处理失败")
            self.log_message.emit(f"处理任务失败: {str(e)}")
    
    def export_to_pdf(self, sheet_name):
        """直接导出为PDF文件"""
        try:
            import win32com.client
            
            # 创建Excel应用程序对象
            xlApp = win32com.client.Dispatch("Excel.Application")
            xlApp.Visible = False
            xlApp.DisplayAlerts = False
            
            # 打开工作簿
            workbook = xlApp.Workbooks.Open(self.file_path)
            
            # 选择指定的工作表
            worksheet = workbook.Worksheets(sheet_name)
            worksheet.Activate()
            
            # 确定输出目录
            if self.output_dir and os.path.exists(self.output_dir):
                output_directory = self.output_dir
            else:
                output_directory = os.path.dirname(self.file_path)
            
            # 生成PDF文件名
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename = os.path.join(
                output_directory, 
                f"{base_name}_{sheet_name}_{timestamp}.pdf"
            )
            
            # 导出为PDF
            worksheet.ExportAsFixedFormat(0, pdf_filename)  # 0 表示PDF格式
            
            # 关闭工作簿和应用程序
            workbook.Close(SaveChanges=False)
            xlApp.Quit()
            
            return pdf_filename
            
        except Exception as e:
            # 确保Excel进程被清理
            try:
                if 'xlApp' in locals():
                    xlApp.Quit()
            except:
                pass
            raise e
    
    def print_to_physical_printer(self, sheet_name):
        """打印到物理打印机"""
        try:
            import win32com.client
            import time
            
            # 创建Excel应用程序对象
            xlApp = win32com.client.Dispatch("Excel.Application")
            xlApp.Visible = False
            xlApp.DisplayAlerts = False
            
            # 打开工作簿
            workbook = xlApp.Workbooks.Open(self.file_path)
            
            # 选择指定的工作表
            worksheet = workbook.Worksheets(sheet_name)
            worksheet.Activate()
            
            # 设置打印机
            xlApp.ActivePrinter = self.printer_name
            
            # 执行打印
            worksheet.PrintOut()
            
            # 关闭工作簿和应用程序
            workbook.Close(SaveChanges=False)
            xlApp.Quit()
            
            # 给一些时间让打印完成
            time.sleep(1)
        except Exception as e:
            # 确保Excel进程被清理
            try:
                if 'xlApp' in locals():
                    xlApp.Quit()
            except:
                pass
            raise e

class ExcelPrinterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.sheet_names = []
        self.output_dir = None
        self.init_ui()
        self.load_printers()
        
    def init_ui(self):
        self.setWindowTitle("Excel工作表PDF导出工具")
        self.setGeometry(100, 100, 900, 700)
        
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # 文件选择区域
        file_group = QGroupBox("选择Excel文件")
        file_layout = QHBoxLayout(file_group)
        
        self.file_label = QLabel("未选择文件")
        self.file_label.setWordWrap(True)
        file_button = QPushButton("选择Excel文件")
        file_button.clicked.connect(self.select_file)
        
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(file_button)
        
        # 工作表选择区域
        sheet_group = QGroupBox("工作表选择")
        sheet_layout = QVBoxLayout(sheet_group)
        
        # 全选按钮
        select_layout = QHBoxLayout()
        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        select_layout.addWidget(self.select_all_checkbox)
        select_layout.addStretch()
        
        # 工作表列表
        self.sheet_list = QListWidget()
        self.sheet_list.setSelectionMode(QListWidget.MultiSelection)
        self.sheet_list.itemSelectionChanged.connect(self.update_select_all_state)
        
        sheet_layout.addLayout(select_layout)
        sheet_layout.addWidget(self.sheet_list)
        
        # 输出设置区域
        output_group = QGroupBox("输出设置")
        output_layout = QVBoxLayout(output_group)
        
        # 打印机选择
        printer_layout = QHBoxLayout()
        printer_layout.addWidget(QLabel("选择打印机:"))
        self.printer_combo = QComboBox()
        printer_layout.addWidget(self.printer_combo)
        output_layout.addLayout(printer_layout)
        
        # PDF输出目录选择
        self.pdf_options_layout = QHBoxLayout()
        self.pdf_options_layout.addWidget(QLabel("PDF输出目录:"))
        self.output_dir_label = QLabel("与Excel文件同目录")
        self.output_dir_label.setWordWrap(True)
        self.pdf_options_layout.addWidget(self.output_dir_label)
        self.output_dir_button = QPushButton("选择目录")
        self.output_dir_button.clicked.connect(self.select_output_dir)
        self.pdf_options_layout.addWidget(self.output_dir_button)
        output_layout.addLayout(self.pdf_options_layout)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        self.print_selected_button = QPushButton("导出选中")
        self.print_selected_button.clicked.connect(self.print_selected)
        self.print_selected_button.setEnabled(False)
        
        self.print_all_button = QPushButton("导出全部")
        self.print_all_button.clicked.connect(self.print_all)
        self.print_all_button.setEnabled(False)
        
        self.cancel_button = QPushButton("取消")
        self.cancel_button.clicked.connect(self.cancel_print)
        
        self.open_folder_button = QPushButton("打开输出目录")
        self.open_folder_button.clicked.connect(self.open_output_folder)
        self.open_folder_button.setEnabled(False)
        
        button_layout.addWidget(self.print_selected_button)
        button_layout.addWidget(self.print_all_button)
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.open_folder_button)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)
        
        # 状态显示区域
        status_group = QGroupBox("处理状态")
        status_layout = QHBoxLayout(status_group)
        
        status_layout.addWidget(QLabel("当前状态:"))
        self.status_label = QLabel("未开始")
        self.status_label.setStyleSheet("color: black; font-weight: bold;")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        
        # 日志显示区域
        log_group = QGroupBox("操作日志")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        
        # 添加所有组件到主布局
        main_layout.addWidget(file_group)
        main_layout.addWidget(sheet_group)
        main_layout.addWidget(output_group)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.progress_bar)
        main_layout.addWidget(status_group)
        main_layout.addWidget(log_group)
        
        # 连接打印机选择变化信号
        self.printer_combo.currentTextChanged.connect(self.on_printer_changed)
        
        # 初始化日志
        self.log("程序启动成功")
        
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "选择Excel文件", 
            "", 
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(file_path)
            self.load_sheets()
            self.log(f"已选择文件: {file_path}")
            
    def load_sheets(self):
        try:
            if self.file_path and os.path.exists(self.file_path):
                workbook = load_workbook(self.file_path)
                self.sheet_names = workbook.sheetnames
                workbook.close()
                
                self.sheet_list.clear()
                for sheet_name in self.sheet_names:
                    self.sheet_list.addItem(sheet_name)
                    
                self.print_selected_button.setEnabled(True)
                self.print_all_button.setEnabled(True)
                self.log(f"加载工作表成功，共 {len(self.sheet_names)} 个工作表")
            else:
                self.log("文件不存在")
        except Exception as e:
            self.log(f"加载工作表失败: {str(e)}")
            QMessageBox.warning(self, "错误", f"无法读取Excel文件: {str(e)}")
            
    def toggle_select_all(self, state):
        if state == Qt.Checked:
            self.sheet_list.selectAll()
        else:
            self.sheet_list.clearSelection()
            
    def update_select_all_state(self):
        # 检查是否所有项都被选中
        total_items = self.sheet_list.count()
        selected_items = len(self.sheet_list.selectedItems())
        
        if total_items > 0 and selected_items == total_items:
            self.select_all_checkbox.setChecked(True)
        else:
            self.select_all_checkbox.setChecked(False)
            
    def load_printers(self):
        try:
            # 获取所有可用打印机
            printers = [printer[2] for printer in win32print.EnumPrinters(2)]
            self.printer_combo.addItems(printers)
            
            # 设置默认打印机
            default_printer = win32print.GetDefaultPrinter()
            try:
                index = printers.index(default_printer)
                self.printer_combo.setCurrentIndex(index)
            except ValueError:
                pass
                
            self.log(f"找到 {len(printers)} 台打印机")
        except Exception as e:
            self.log(f"加载打印机失败: {str(e)}")
            
    def on_printer_changed(self, printer_name):
        """当打印机选择改变时调用"""
        # 检查是否选择了PDF打印机
        is_pdf_printer = "Adobe PDF" in printer_name or "Microsoft Print to PDF" in printer_name
        # 显示或隐藏PDF选项
        for i in range(self.pdf_options_layout.count()):
            widget = self.pdf_options_layout.itemAt(i).widget()
            if widget:
                widget.setVisible(is_pdf_printer)
                
    def select_output_dir(self):
        """选择PDF输出目录"""
        directory = QFileDialog.getExistingDirectory(self, "选择PDF输出目录")
        if directory:
            self.output_dir = directory
            self.output_dir_label.setText(directory)
            self.log(f"PDF输出目录设置为: {directory}")
            
    def print_selected(self):
        selected_items = self.sheet_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请至少选择一个工作表")
            return
            
        sheet_names = [item.text() for item in selected_items]
        self.start_printing(sheet_names)
        
    def print_all(self):
        if not self.sheet_names:
            QMessageBox.warning(self, "警告", "没有可处理的工作表")
            return
            
        self.start_printing(self.sheet_names)
        
    def start_printing(self, sheet_names):
        if not self.file_path:
            QMessageBox.warning(self, "警告", "请先选择Excel文件")
            return
            
        printer_name = self.printer_combo.currentText()
        if not printer_name:
            QMessageBox.warning(self, "警告", "请选择打印机")
            return
            
        # 禁用按钮防止重复点击
        self.print_selected_button.setEnabled(False)
        self.print_all_button.setEnabled(False)
        self.select_file_button_state(False)
        self.open_folder_button.setEnabled(False)
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 创建并启动处理线程
        self.printer_worker = PrinterWorker(self.file_path, sheet_names, printer_name, self.output_dir)
        self.printer_worker.print_status.connect(self.update_status)
        self.printer_worker.log_message.connect(self.log)
        self.printer_worker.progress_update.connect(self.update_progress)
        self.printer_worker.finished.connect(self.print_finished)
        self.printer_worker.start()
        
    def select_file_button_state(self, enabled):
        # 控制文件选择按钮状态
        for button in self.findChildren(QPushButton):
            if "选择Excel文件" in button.text():
                button.setEnabled(enabled)
                break
                
    def update_status(self, status):
        self.status_label.setText(status)
        if status == "处理成功":
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
        elif status == "处理失败":
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
        elif status == "正在处理":
            self.status_label.setStyleSheet("color: blue; font-weight: bold;")
        else:
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            
    def update_progress(self, value):
        self.progress_bar.setValue(value)
            
    def log(self, message):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
        
    def print_finished(self):
        # 处理完成后重新启用按钮
        self.print_selected_button.setEnabled(True)
        self.print_all_button.setEnabled(True)
        self.select_file_button_state(True)
        
        # 启用打开文件夹按钮
        self.open_folder_button.setEnabled(True)
        
        # 隐藏进度条
        self.progress_bar.setVisible(False)
        
        self.log("处理任务完成")
        
    def cancel_print(self):
        if hasattr(self, 'printer_worker') and self.printer_worker.isRunning():
            self.printer_worker.terminate()
            self.printer_worker.wait()
            self.update_status("已取消")
            self.log("处理任务已取消")
            self.progress_bar.setVisible(False)
        else:
            self.log("没有正在运行的处理任务")
            
    def open_output_folder(self):
        """打开输出文件夹"""
        try:
            # 确定要打开的目录
            if self.output_dir and os.path.exists(self.output_dir):
                folder_to_open = self.output_dir
            elif self.file_path:
                folder_to_open = os.path.dirname(self.file_path)
            else:
                folder_to_open = os.path.expanduser("~")  # 用户主目录
            
            # 打开文件夹
            if os.name == 'nt':  # Windows
                os.startfile(folder_to_open)
            elif os.name == 'posix':  # macOS or Linux
                subprocess.Popen(['open', folder_to_open])
                
            self.log(f"已打开文件夹: {folder_to_open}")
        except Exception as e:
            self.log(f"打开文件夹失败: {str(e)}")
            QMessageBox.warning(self, "错误", f"无法打开文件夹: {str(e)}")

def main():
    app = QApplication(sys.argv)
    window = ExcelPrinterApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()