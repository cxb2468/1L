import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QListWidget, QComboBox, QLabel, QFileDialog,
    QTextEdit, QCheckBox, QMessageBox, QGroupBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from openpyxl import load_workbook
import win32print

class PrinterWorker(QThread):
    # 定义信号，用于线程与主线程通信
    print_status = pyqtSignal(str)  # 打印状态信号
    log_message = pyqtSignal(str)   # 日志信息信号
    
    def __init__(self, file_path, sheet_names, printer_name):
        super().__init__()
        self.file_path = file_path
        self.sheet_names = sheet_names
        self.printer_name = printer_name
        
    def run(self):
        try:
            self.log_message.emit(f"开始打印任务: {self.file_path}")
            
            # 如果没有指定工作表，则打印所有工作表
            if not self.sheet_names:
                workbook = load_workbook(self.file_path)
                self.sheet_names = workbook.sheetnames
                workbook.close()
            
            self.log_message.emit(f"将要打印的工作表: {', '.join(self.sheet_names)}")
            
            # 模拟打印过程
            for sheet_name in self.sheet_names:
                self.print_status.emit("正在打印")
                self.log_message.emit(f"正在处理工作表: {sheet_name}")
                
                # 模拟打印耗时
                import time
                time.sleep(2)
                
                # 这里应该实现实际的打印逻辑
                # 由于真实的打印需要复杂的COM组件交互，我们在此仅模拟
                try:
                    # 模拟打印成功
                    self.print_status.emit("打印成功")
                    self.log_message.emit(f"工作表 {sheet_name} 打印完成")
                except Exception as e:
                    self.print_status.emit("打印失败")
                    self.log_message.emit(f"工作表 {sheet_name} 打印失败: {str(e)}")
                    
        except Exception as e:
            self.print_status.emit("打印失败")
            self.log_message.emit(f"打印任务失败: {str(e)}")

class ExcelPrinterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.sheet_names = []
        self.init_ui()
        self.load_printers()
        
    def init_ui(self):
        self.setWindowTitle("Excel工作表打印工具")
        self.setGeometry(100, 100, 800, 600)
        
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # 文件选择区域
        file_group = QGroupBox("选择Excel文件")
        file_layout = QHBoxLayout(file_group)
        
        self.file_label = QLabel("未选择文件")
        self.file_label.setWordWrap(True)
        file_button = QPushButton("选择Excel文件")
        file_button.clicked.connect(self.select_file)
        
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(file_button)
        
        # 工作表选择区域
        sheet_group = QGroupBox("工作表选择")
        sheet_layout = QVBoxLayout(sheet_group)
        
        # 全选按钮
        select_layout = QHBoxLayout()
        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        select_layout.addWidget(self.select_all_checkbox)
        select_layout.addStretch()
        
        # 工作表列表
        self.sheet_list = QListWidget()
        self.sheet_list.setSelectionMode(QListWidget.MultiSelection)
        self.sheet_list.itemSelectionChanged.connect(self.update_select_all_state)
        
        sheet_layout.addLayout(select_layout)
        sheet_layout.addWidget(self.sheet_list)
        
        # 打印机选择区域
        printer_group = QGroupBox("打印机选择")
        printer_layout = QHBoxLayout(printer_group)
        
        printer_layout.addWidget(QLabel("选择打印机:"))
        self.printer_combo = QComboBox()
        printer_layout.addWidget(self.printer_combo)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        self.print_selected_button = QPushButton("打印选中")
        self.print_selected_button.clicked.connect(self.print_selected)
        self.print_selected_button.setEnabled(False)
        
        self.print_all_button = QPushButton("打印全部")
        self.print_all_button.clicked.connect(self.print_all)
        self.print_all_button.setEnabled(False)
        
        self.cancel_button = QPushButton("取消打印")
        self.cancel_button.clicked.connect(self.cancel_print)
        
        button_layout.addWidget(self.print_selected_button)
        button_layout.addWidget(self.print_all_button)
        button_layout.addWidget(self.cancel_button)
        
        # 状态显示区域
        status_group = QGroupBox("打印状态")
        status_layout = QHBoxLayout(status_group)
        
        status_layout.addWidget(QLabel("当前状态:"))
        self.status_label = QLabel("未开始")
        self.status_label.setStyleSheet("color: black; font-weight: bold;")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        
        # 日志显示区域
        log_group = QGroupBox("操作日志")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        
        # 添加所有组件到主布局
        main_layout.addWidget(file_group)
        main_layout.addWidget(sheet_group)
        main_layout.addWidget(printer_group)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(status_group)
        main_layout.addWidget(log_group)
        
        # 初始化日志
        self.log("程序启动成功")
        
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "选择Excel文件", 
            "", 
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(file_path)
            self.load_sheets()
            self.log(f"已选择文件: {file_path}")
            
    def load_sheets(self):
        try:
            if self.file_path and os.path.exists(self.file_path):
                workbook = load_workbook(self.file_path)
                self.sheet_names = workbook.sheetnames
                workbook.close()
                
                self.sheet_list.clear()
                for sheet_name in self.sheet_names:
                    self.sheet_list.addItem(sheet_name)
                    
                self.print_selected_button.setEnabled(True)
                self.print_all_button.setEnabled(True)
                self.log(f"加载工作表成功，共 {len(self.sheet_names)} 个工作表")
            else:
                self.log("文件不存在")
        except Exception as e:
            self.log(f"加载工作表失败: {str(e)}")
            QMessageBox.warning(self, "错误", f"无法读取Excel文件: {str(e)}")
            
    def toggle_select_all(self, state):
        if state == Qt.Checked:
            self.sheet_list.selectAll()
        else:
            self.sheet_list.clearSelection()
            
    def update_select_all_state(self):
        # 检查是否所有项都被选中
        total_items = self.sheet_list.count()
        selected_items = len(self.sheet_list.selectedItems())
        
        if total_items > 0 and selected_items == total_items:
            self.select_all_checkbox.setChecked(True)
        else:
            self.select_all_checkbox.setChecked(False)
            
    def load_printers(self):
        try:
            # 获取所有可用打印机
            printers = [printer[2] for printer in win32print.EnumPrinters(2)]
            self.printer_combo.addItems(printers)
            
            # 设置默认打印机
            default_printer = win32print.GetDefaultPrinter()
            try:
                index = printers.index(default_printer)
                self.printer_combo.setCurrentIndex(index)
            except ValueError:
                pass
                
            self.log(f"找到 {len(printers)} 台打印机")
        except Exception as e:
            self.log(f"加载打印机失败: {str(e)}")
            
    def print_selected(self):
        selected_items = self.sheet_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请至少选择一个工作表")
            return
            
        sheet_names = [item.text() for item in selected_items]
        self.start_printing(sheet_names)
        
    def print_all(self):
        if not self.sheet_names:
            QMessageBox.warning(self, "警告", "没有可打印的工作表")
            return
            
        self.start_printing(self.sheet_names)
        
    def start_printing(self, sheet_names):
        if not self.file_path:
            QMessageBox.warning(self, "警告", "请先选择Excel文件")
            return
            
        printer_name = self.printer_combo.currentText()
        if not printer_name:
            QMessageBox.warning(self, "警告", "请选择打印机")
            return
            
        # 禁用按钮防止重复点击
        self.print_selected_button.setEnabled(False)
        self.print_all_button.setEnabled(False)
        self.select_file_button_state(False)
        
        # 创建并启动打印线程
        self.printer_worker = PrinterWorker(self.file_path, sheet_names, printer_name)
        self.printer_worker.print_status.connect(self.update_status)
        self.printer_worker.log_message.connect(self.log)
        self.printer_worker.finished.connect(self.print_finished)
        self.printer_worker.start()
        
    def select_file_button_state(self, enabled):
        # 控制文件选择按钮状态
        for button in self.findChildren(QPushButton):
            if "选择Excel文件" in button.text():
                button.setEnabled(enabled)
                break
                
    def update_status(self, status):
        self.status_label.setText(status)
        if status == "打印成功":
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
        elif status == "打印失败":
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
        elif status == "正在打印":
            self.status_label.setStyleSheet("color: blue; font-weight: bold;")
        else:
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            
    def log(self, message):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
        
    def print_finished(self):
        # 打印完成后重新启用按钮
        self.print_selected_button.setEnabled(True)
        self.print_all_button.setEnabled(True)
        self.select_file_button_state(True)
        self.log("打印任务完成")
        
    def cancel_print(self):
        if hasattr(self, 'printer_worker') and self.printer_worker.isRunning():
            self.printer_worker.terminate()
            self.printer_worker.wait()
            self.update_status("已取消")
            self.log("打印任务已取消")
        else:
            self.log("没有正在运行的打印任务")

def main():
    app = QApplication(sys.argv)
    window = ExcelPrinterApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()