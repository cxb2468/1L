# -*- coding: utf-8 -*-

import win32com.client.gencache as gencache
import re
from win32com.client.gencache import EnsureDispatch as Dispatch  # 读取邮件模块
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
import win32timezone


def read_outlook_mailbox(specific_time):
    """连接Outlook邮箱读取收件箱内的邮件内容"""
    # 使用MAPI协议连接Outlook
    account = gencache.EnsureDispatch('Outlook.Application').GetNamespace('MAPI')
    print("Outlook Namespace:", account)
    # 获取收件箱所在位置
    inbox = account.GetDefaultFolder(6)  # 数字6代表收件箱
    # 获取收件箱下的所有邮件

    start_time = specific_time
    # 读取范围为指定日期之后一周内的邮件
    end_time = start_time + timedelta(days=360)
    filtered_mails = inbox.Items
    filtered_mails.Sort('[ReceivedTime]', True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "location"

    f = open("email_location_log.txt", "w+", encoding="utf-8")
    number_test_floor = 0
    # 读取收件箱内邮件的所有信息（下标从1开始）
    for mail in filtered_mails:
        # 找到所有范围内的邮箱中，标题带有"TEST FLOOR"的邮件
        pattern_test_floor = re.compile(r'域控服务器补丁更新时间确认', re.IGNORECASE)

        if pattern_test_floor.search(mail.Subject):

            mail_time = str(mail.ReceivedTime)[:19]
            mail_true_time = datetime.strptime(mail_time, '%Y-%m-%d %H:%M:%S')
            if (mail_true_time >= start_time) and (mail_true_time <= end_time):
                number_test_floor += 1
                temp_list = []
                # 将邮件信息写入到文件中
                f.write('正在读取{}第[{}]封邮件...'.format("TESTFLOOR", number_test_floor))
                f.write("\n")
                f.write('接收时间：{}'.format(str(mail.ReceivedTime)[:19]))
                f.write('\n')
                f.write('发件人：{}'.format(mail.SenderName))
                f.write('\n')
                f.write('主题：{}'.format(mail.Subject))
                f.write('\n')
                f.write('邮件正文内容：{}'.format(mail.Body))
                f.write('\n')
                temp_rule = re.compile(r'X0(?!.*域控服务器补丁更新时间确认).*')
                temp_content = temp_rule.findall(mail.Body)

                # temp_content = [words for words in temp_content if 'TEST FLOOR' not in word]
                #  it=1
                #  for temp_word in temp_content:
                #     print("list {}={}".format(it,temp_word))
                #     it+=1

                sheet.cell(row=1, column=number_test_floor, value=str(mail.ReceivedTime)[:16])
                # sheet.column_dimensions[number_test_floor].width=15
                sheet.column_dimensions[get_column_letter(number_test_floor)].width = 20
                j = 2
                for word in temp_content:
                    sheet.cell(row=j, column=number_test_floor, value=word)
                    j += 1

    now_max_rows = sheet.max_row  # 获取行数
    now_max_rows += 3
    temp_rows = now_max_rows

    # print("max rows1=",rows)
    # print("max columns1=",column)
    number_burnin = 0
    for mail in filtered_mails:
        pattern_burnin = re.compile(r'BURNIN', re.IGNORECASE)

        if pattern_burnin.search(mail.Subject):
            mail_time = str(mail.ReceivedTime)[:19]
            mail_true_time = datetime.strptime(mail_time, '%Y-%m-%d %H:%M:%S')
            if (mail_true_time >= start_time) and (mail_true_time <= end_time):
                number_burnin += 1
                # 将邮件信息写入到文件中
                f.write('正在读取{}第[{}]封邮件...'.format("BURNIN", number_burnin))
                f.write("\n")
                f.write('接收时间：{}'.format(str(mail.ReceivedTime)[:19]))
                f.write('\n')
                f.write('发件人：{}'.format(mail.SenderName))
                f.write('\n')

                f.write('主题：{}'.format(mail.Subject))
                f.write('\n')
                f.write('邮件正文内容：{}'.format(mail.Body))
                f.write('\n')
                temp_rule = re.compile(r'X0(?!.*BURNIN).*')
                temp_content = temp_rule.findall(mail.Body)
                it = 1
                for temp_word in temp_content:
                    print("list {}={}".format(it, temp_word))
                    it += 1
                sheet.cell(row=temp_rows, column=number_burnin, value=str(mail.ReceivedTime)[:16])
                # sheet.column_dimensions[number_test_floor].width=15
                sheet.column_dimensions[get_column_letter(number_burnin)].width = 20
                j = temp_rows + 1
                for word in temp_content:
                    sheet.cell(row=j, column=number_burnin, value=word)
                    j += 1

    now_max_rows = sheet.max_row  # 获取行数
    now_max_rows += 3
    temp_rows = now_max_rows

    # print("max rows1=",rows)
    # print("max columns1=",column)
    number_PE = 0
    for mail in filtered_mails:
        pattern_burnin = re.compile(r'POST ELECTRICAL', re.IGNORECASE)

        if pattern_burnin.search(mail.Subject):
            mail_time = str(mail.ReceivedTime)[:19]
            mail_true_time = datetime.strptime(mail_time, '%Y-%m-%d %H:%M:%S')
            if (mail_true_time >= start_time) and (mail_true_time <= end_time):
                number_PE += 1
                # 将邮件信息写入到文件中
                f.write('正在读取{}第[{}]封邮件...'.format("PE", number_PE))
                f.write("\n")
                f.write('接收时间：{}'.format(str(mail.ReceivedTime)[:19]))
                f.write('\n')
                f.write('发件人：{}'.format(mail.SenderName))
                f.write('\n')

                f.write('主题：{}'.format(mail.Subject))
                f.write('\n')
                f.write('邮件正文内容：{}'.format(mail.Body))
                f.write('\n')
                temp_rule = re.compile(r'X0(?!.*POST ELECTRICAL).*')
                temp_content = temp_rule.findall(mail.Body)
                it = 1
                for temp_word in temp_content:
                    print("list {}={}".format(it, temp_word))
                    it += 1
                sheet.cell(row=temp_rows, column=number_PE, value=str(mail.ReceivedTime)[:16])
                # sheet.column_dimensions[number_test_floor].width=15
                sheet.column_dimensions[get_column_letter(number_PE)].width = 20
                j = temp_rows + 1
                for word in temp_content:
                    sheet.cell(row=j, column=number_PE, value=word)
                    j += 1

    workbook.save("location_excel.xlsx")


user_input_time = input("please input date(formation:DD/MM/YYYY):")
start_date = datetime.strptime(user_input_time, '%d/%m/%Y')
specific_time = datetime(start_date.year, start_date.month, start_date.day, 19, 00, 0)

read_outlook_mailbox(specific_time)