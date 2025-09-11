import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import os

# 使用openpyxl进行保存并设置列宽
def save_df_to_excel_with_column_width(df, output_file_path, column_widths):
    """Save DataFrame to Excel with specified column widths, cell alignment, and borders."""
    wb = Workbook()
    ws = wb.active

    # 添加标题行
    title_row = ['用户管理']
    for col in range(2, 10):
        title_row.append(None)
    ws.append(title_row)

    # 合并单元格并设置标题样式
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = 'U8用户管理'
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_cell.font = Font(bold=True, size=20)

    # 写入DataFrame数据
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Set column width
    for idx, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    # Set cell alignment to left
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add border to cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    # Add bold font to header row
    for cell in ws[2]:  # ws[2] refers to the second row (header row after title row)
        cell.font = Font(bold=True)

    try:
        wb.save(output_file_path)
    except PermissionError:
        print("文件正在被其他程序使用，请关闭后再试。")
    except Exception as e:
        print(f"An error occurred while saving the file: {e}")

# Excel处理函数
# Update process_excel to use the improved save function
def process_excel(input_file_path, output_file_path):
    if os.path.exists(output_file_path):
        response = tk.messagebox.askyesno("警告", "输出文件已存在，是否覆盖？")
        if not response:
            return

    df = pd.read_excel(input_file_path)
    column_order = [
        '用户编码', '用户全名', '用户类型', '认证方式',
        '状态', '最后登录时间', '退出时间', '权限变更'
    ]
    df['权限变更'] = ''  # 添加新列，数据为空
    df_reordered = df[column_order]
    df_reordered.insert(0, '序号', np.arange(1, len(df_reordered) + 1))

    column_widths = [5, 9, 9, 9, 9, 5, 20, 20, 16]  # 增加新列的宽度
    save_df_to_excel_with_column_width(df_reordered, output_file_path, column_widths)

    print(f"Excel文件处理完成，保存路径为：{output_file_path}")

# 选择文件对话框
def select_input_file(entry):
    filepath = filedialog.askopenfilename()
    entry.delete(0, tk.END)
    entry.insert(0, filepath)

# 选择输出目录对话框
def select_output_directory(entry):
    directory = filedialog.askdirectory()
    output_path = os.path.join(directory, "9.2.1 U8用户管理.xlsx")
    # Normalize the path or replace backslashes for display purposes
    normalized_path = os.path.normpath(output_path).replace("\\", "/")
    entry.delete(0, tk.END)
    entry.insert(0, normalized_path)

# GUI提交处理逻辑
def submit(input_entry, output_entry, submit_button):
    input_file_path = input_entry.get()
    output_file_path = output_entry.get()
    try:
        process_excel(input_file_path, output_file_path)
        tk.messagebox.showinfo("完成", "Excel文件处理完成！")
    except FileNotFoundError:
        tk.messagebox.showerror("错误", "输入的文件未找到，请检查路径。")
    except Exception as e:
        tk.messagebox.showerror("错误", f"处理过程中发生错误：{str(e)}")
    finally:
        submit_button.config(text="处理完成", state=tk.DISABLED)

# GUI设置
def setup_gui():
    global root
    root = tk.Tk()
    root.title("Excel处理工具(用友U8)")
    root.geometry("350x500")
    root.configure(bg="#F0F0F0")
    root.resizable(True, True)

    # 省略具体组件定义，与您提供的GUI代码相同...
    tk.Label(root, text="选择输入文件:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    input_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    input_entry.pack(ipady=4, pady=10, padx=10)
    tk.Button(root, text="浏览", command=lambda: select_input_file(input_entry), font=("Arial", 12), bg="#4CAF50",
              fg="white", activebackground="#45a049", relief=tk.FLAT, cursor="hand2").pack(pady=5, padx=10)

    tk.Label(root, text="选择输出目录:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    output_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    output_entry.pack(ipady=4, pady=10, padx=10)
    tk.Button(root, text="浏览", command=lambda: select_output_directory(output_entry), font=("Arial", 12), bg="#4CAF50",
              fg="white", activebackground="#45a049", relief=tk.FLAT, cursor="hand2").pack(pady=5, padx=10)

    # tk.Label(root, text="输入列索引:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    # column_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    # column_entry.pack(ipady=4, pady=10, padx=10)

    # 修改submit_button的command参数，以调用整合后的处理逻辑
    submit_button = tk.Button(root, text="开始处理",
                              command=lambda: submit(input_entry, output_entry, submit_button),
                              font=("Arial", 12), bg="#4CAF50", fg="white", activebackground="#45a049", relief=tk.FLAT,
                              cursor="hand2")
    submit_button.pack(pady=20, padx=10)

    root.mainloop()

if __name__ == "__main__":
    # 启动GUI
    setup_gui()
