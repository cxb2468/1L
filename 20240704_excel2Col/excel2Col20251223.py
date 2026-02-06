import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import os



# 使用openpyxl进行保存并设置列宽
def save_df_to_excel_with_column_width(df, output_file_path, column_widths=None):
    """Save DataFrame to Excel with specified column widths and cell alignment."""
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Set column width dynamically if column_widths is not provided
    if column_widths is None:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    else:
        for idx, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

    # Set cell alignment to left
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
            cell.border = thin_border

    # Add bold font to header row
    for cell in ws[1]:  # ws[1] refers to the first row
        cell.font = Font(bold=True)


    try:
        wb.save(output_file_path)
    except PermissionError:
        print("文件正在被其他程序使用，请关闭后再试。")
    except Exception as e:
        print(f"An error occurred while saving the file: {e}")


# 假设 pinyin_first_letter 是一个字典，键为汉字，值为拼音首字母
pinyin_first_letter = {
    '部': 'b',
    '门': 'm',
    '卫': 'w',
    '资': 'z',
    '材': 'c',
    '课': 'k',
    '财': 'c',
    '务': 'w',
    '管': 'g',
    '理': 'l',
    '生': 's',
    '环': 'h',
    '境': 'j',
    '安': 'a',
    '全': 'q',
    '物': 'w',
    '技': 'j',
    '术': 's',
    '开': 'k',
    '发': 'f',
    '总': 'z',
    '经': 'j',
    '办': 'b',
    '工': 'g',
    '程': 'c',
    '设': 's',
    '备': 'b',
    '品': 'p',
    '质': 'z',
    '保': 'b',
    '证': 'z',
    '人': 'r',
    '力': 'l',
    '源': 'y',
    '二': 'e',
    '产': 'c',
    '室': 's',
    '流': 'l',
}

# 使用该函数创建一个新的列“部门首字母”
def get_pinyin_initials(department):
    return ''.join(pinyin_first_letter.get(char, char) for char in department)


# Excel处理函数
def process_excel(input_file_path, output_file_path):
    if os.path.exists(output_file_path):
        response = tk.messagebox.askyesno("警告", "输出文件已存在，是否覆盖？")
        if not response:
            return

    df = pd.read_excel(input_file_path)
    column_order = [
        '部门', '用户', '名称', '厂商','型号',
        'CPU型号', '主频', '内存(G)',
        '操作系统', '软件', '购入日期', '分类', '用途', '固定资产编号'
    ]
    df_reordered = df[column_order].copy()  # 创建一个副本以避免潜在的SettingWithCopyWarning

    # 使用 get_pinyin_initials 函数创建一个新的列“部门首字母”
    df_reordered['部门首字母'] = df_reordered['部门'].apply(get_pinyin_initials)

    # 根据“部门首字母”列对 DataFrame 进行排序
    df_sorted = df_reordered.sort_values(by=['部门首字母', '用户'], ascending=(False, False))

    # 删除“部门首字母”列
    df_sorted.drop(columns=['部门首字母'], inplace=True)

    # 在排序后的 DataFrame 上插入序号列
    df_sorted.insert(0, '序号', np.arange(1, len(df_sorted) + 1))

    column_widths = [6, 12, 12, 15, 12, 17, 12, 12, 12, 12, 52, 12, 12, 15, 15]
    # 使用排序后的 DataFrame 保存文件
    save_df_to_excel_with_column_width(df_sorted, output_file_path, column_widths)

    print(f"\nExcel文件处理完成，保存路径为：{output_file_path}")


# 选择文件对话框
def select_input_file(entry):
    filepath = filedialog.askopenfilename()
    entry.delete(0, tk.END)
    entry.insert(0, filepath)


# 选择输出目录对话框
def select_output_directory(entry):
    directory = filedialog.askdirectory()
    output_path = os.path.join(directory, "电脑管理台账.xlsx")
    # Normalize the path or replace backslashes for display purposes
    normalized_path = os.path.normpath(output_path).replace("\\", "/")
    entry.delete(0, tk.END)
    entry.insert(0, normalized_path)


# GUI提交处理逻辑
def submit(input_entry, output_entry, submit_button):
    input_file_path = input_entry.get()
    output_file_path = output_entry.get()

    # 检查文件路径是否为空
    if not input_file_path or not output_file_path:
        tk.messagebox.showerror("错误", "请输入完整的文件路径。")
        return

    try:
        process_excel(input_file_path, output_file_path)
        tk.messagebox.showinfo("完成", "Excel文件处理完成！")
    except FileNotFoundError:
        tk.messagebox.showerror("错误", "输入的文件未找到，请检查路径。")
    except Exception as e:
        tk.messagebox.showerror("错误", f"处理过程中发生错误：{str(e)}")
    finally:
        submit_button.config(text="处理完成", state=tk.DISABLED)


# GUI设置
def setup_gui():
    global root
    root = tk.Tk()
    root.title("Excel处理工具(IT资产管理系统)")
    root.geometry("350x500")
    root.configure(bg="#F0F0F0")
    root.resizable(True, True)

    # 省略具体组件定义，与您提供的GUI代码相同...
    tk.Label(root, text="选择输入文件:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    input_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    input_entry.pack(ipady=4, pady=10, padx=10)
    tk.Button(root, text="浏览", command=lambda: select_input_file(input_entry), font=("Arial", 12), bg="#4CAF50",
              fg="white", activebackground="#45a049", relief=tk.FLAT, cursor="hand2").pack(pady=5, padx=10)

    tk.Label(root, text="选择输出目录:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    output_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    output_entry.pack(ipady=4, pady=10, padx=10)
    tk.Button(root, text="浏览", command=lambda: select_output_directory(output_entry), font=("Arial", 12), bg="#4CAF50",
              fg="white", activebackground="#45a049", relief=tk.FLAT, cursor="hand2").pack(pady=5, padx=10)

    # tk.Label(root, text="输入列索引:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    # column_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    # column_entry.pack(ipady=4, pady=10, padx=10)

    # 修改submit_button的command参数，以调用整合后的处理逻辑
    submit_button = tk.Button(root, text="开始处理",
                              command=lambda: submit(input_entry, output_entry, submit_button),
                              font=("Arial", 12), bg="#4CAF50", fg="white", activebackground="#45a049", relief=tk.FLAT,
                              cursor="hand2")
    submit_button.pack(pady=20, padx=10)

    root.mainloop()


if __name__ == "__main__":
    # 启动GUI
    setup_gui()
