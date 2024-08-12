import tkinter as tk
from tkinter import filedialog, messagebox
import os
import threading
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def collect_data_and_style(sheet, key_column_index=0):
    """收集单个工作表的数据及其样式信息，包括列头，现在可以指定分组列，并处理百分比格式"""
    groups = {}
    header_info = []
    for row in sheet.iter_rows(min_row=1):
        if row[0].row == 1:
            header_info = [(cell.value, {
                'font': Font(name=cell.font.name, sz=cell.font.sz, b=cell.font.b, i=cell.font.i,
                             u=cell.font.u, strike=cell.font.strike, color=cell.font.color),
                'alignment': Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical,
                                       wrapText=cell.alignment.wrapText, shrinkToFit=cell.alignment.shrinkToFit),
                'fill': PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color, patternType=cell.fill.patternType),
                'border': Border(left=Side(border_style=cell.border.left.border_style, color=cell.border.left.color),
                                 right=Side(border_style=cell.border.right.border_style, color=cell.border.right.color),
                                 top=Side(border_style=cell.border.top.border_style, color=cell.border.top.color),
                                 bottom=Side(border_style=cell.border.bottom.border_style,
                                             color=cell.border.bottom.color)),
                'number_format': cell.number_format,  # 添加这一行来收集单元格的数字格式
            }) for cell in row]
            continue

        group_key = str(row[key_column_index].value)

        if group_key not in groups:
            groups[group_key] = []
        row_data = [(cell.value, {
            'font': Font(name=cell.font.name, sz=cell.font.sz, b=cell.font.b, i=cell.font.i,
                         u=cell.font.u, strike=cell.font.strike, color=cell.font.color),
            'alignment': Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical,
                                   wrapText=cell.alignment.wrapText, shrinkToFit=cell.alignment.shrinkToFit),
            'fill': PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color,
                                end_color=cell.fill.end_color, patternType=cell.fill.patternType),
            'border': Border(left=Side(border_style=cell.border.left.border_style, color=cell.border.left.color),
                             right=Side(border_style=cell.border.right.border_style, color=cell.border.right.color),
                             top=Side(border_style=cell.border.top.border_style, color=cell.border.top.color),
                             bottom=Side(border_style=cell.border.bottom.border_style, color=cell.border.bottom.color)),
            'number_format': cell.number_format,  # 同样收集每个单元格的数字格式
        }) for cell in row]
        groups[group_key].append(row_data)

    for group_data in groups.values():
        group_data.insert(0, header_info)

    return groups


def create_merged_excel(output_path, grouped_data):
    output_wb = Workbook()
    for sheet_name, sheet_data in grouped_data.items():
        ws = output_wb.create_sheet(title=sheet_name)
        for row_index, row_data_with_styles in enumerate(sheet_data, start=1):
            for col_index, (value, style_info) in enumerate(row_data_with_styles, start=1):
                new_cell = ws.cell(row=row_index, column=col_index, value=value)
                if style_info:
                    new_cell.font = style_info.get('font', None)
                    new_cell.alignment = style_info.get('alignment', None)
                    new_cell.fill = style_info.get('fill', None)
                    new_cell.border = style_info.get('border', None)
                    number_format = style_info.get('number_format', '')
                    if '%)' in number_format:
                        new_cell.value *= 100
                        new_cell.number_format = number_format.replace('0', '0%')
                    else:
                        new_cell.number_format = number_format
    # 删除默认的Sheet，如果有的话
    default_sheet_name = "Sheet"
    if default_sheet_name in output_wb.sheetnames:
        del output_wb[default_sheet_name]
    output_wb.save(output_path)
    output_wb.close()


def extract_and_merge_sheets(input_file, output_directory, key_column_index=0):
    """主函数，负责遍历输入文件的所有工作表，收集数据并创建合并的Excel文件，处理百分比等格式问题"""
    wb = load_workbook(input_file, read_only=True)
    all_groups_data = {}
    for sheet in wb:
        sheet_groups = collect_data_and_style(sheet, key_column_index=key_column_index)
        for group_key, group_data in sheet_groups.items():
            if group_key not in all_groups_data:
                all_groups_data[group_key] = {}
            all_groups_data[group_key][sheet.title] = group_data

    os.makedirs(output_directory, exist_ok=True)
    for group, sheets_data in all_groups_data.items():
        if sheets_data:
            output_path = os.path.join(output_directory, f"{group}.xlsx")
            create_merged_excel(output_path, {sheet_name: data for sheet_name, data in sheets_data.items() if data})


def select_input_file(entry_widget):
    filepath = filedialog.askopenfilename()
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, filepath)


def select_output_directory(entry_widget):
    directory = filedialog.askdirectory()
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, directory)


def process_in_background(input_path, output_directory, column_index, submit_button):
    try:
        submit_button.config(state=tk.DISABLED, bg="#FF0000")
        extract_and_merge_sheets(input_path, output_directory, column_index)
        submit_button.config(state=tk.NORMAL, bg="#4CAF50")
        messagebox.showinfo("完成", "处理结束。")
    except ValueError:
        submit_button.config(state=tk.NORMAL, bg="#4CAF50")
        messagebox.showerror("错误", "列索引必须是整数。")
    except Exception as e:
        submit_button.config(state=tk.NORMAL, bg="#4CAF50")
        messagebox.showerror("错误", f"处理过程中发生错误: {str(e)}")


def submit(input_var, output_var, column_var, submit_button):
    input_path = input_var.get()
    output_directory = output_var.get()
    try:
        column_index = int(column_var.get())
        thread = threading.Thread(target=process_in_background,
                                  args=(input_path, output_directory, column_index, submit_button))
        thread.start()
    except ValueError:
        messagebox.showerror("错误", "列索引必须是整数。")


def setup_gui():
    global root
    root = tk.Tk()
    root.title("Excel拆分工具")
    root.geometry("350x500")
    root.configure(bg="#F0F0F0")
    root.resizable(True, True)

    tk.Label(root, text="选择输入文件:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    input_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    input_entry.pack(ipady=4, pady=10, padx=10)
    tk.Button(root, text="浏览", command=lambda: select_input_file(input_entry), font=("Arial", 12), bg="#4CAF50",
              fg="white", activebackground="#45a049", relief=tk.FLAT, cursor="hand2").pack(pady=5, padx=10)

    tk.Label(root, text="选择输出目录:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    output_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    output_entry.pack(ipady=4, pady=10, padx=10)
    tk.Button(root, text="浏览", command=lambda: select_output_directory(output_entry), font=("Arial", 12), bg="#4CAF50",
              fg="white", activebackground="#45a049", relief=tk.FLAT, cursor="hand2").pack(pady=5, padx=10)

    tk.Label(root, text="输入列索引:", font=("Arial", 12), bg="#F0F0F0").pack(pady=20)
    column_entry = tk.Entry(root, font=("Arial", 12), relief=tk.FLAT, bg="white")
    column_entry.pack(ipady=4, pady=10, padx=10)

    submit_button = tk.Button(root, text="开始处理",
                              command=lambda: submit(input_entry, output_entry, column_entry, submit_button),
                              font=("Arial", 12), bg="#4CAF50", fg="white", activebackground="#45a049", relief=tk.FLAT,
                              cursor="hand2")
    submit_button.pack(pady=20, padx=10)

    root.mainloop()


if __name__ == "__main__":
    setup_gui()