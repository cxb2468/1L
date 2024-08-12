"""
把excel表按照某列拆分成多个表
具体为保留前六列的前提下，把第六列以后的每一列都形成一个分表 ，并且不存在空元素
"""
import excel_read_write
from openpyxl.utils import get_column_letter, column_index_from_string

# openpyxl.load_workbook(需要打开的excel文件路径)
wb = excel_read_write.load_workbook("D:\\1.xlsx")
print(type(wb))  # 结果: <class 'openpyxl.workbook.workbook.Workbook'>

# 获取所有表的表名
sheets_names = wb.sheetnames
print(sheets_names)  # 结果: ['sheet1', 'sheet2']

# 获取活动表对应的表对象(表对象就是Worksheet类的对象)
sheet = wb.active
print(sheet)  # 结果：<Worksheet "sheet1">

# # 根据表名获取工作簿中指定的表
# sheet2 = wb['sheet2']
# print(sheet2)  # 结果：<Worksheet "sheet2">
#
# # 根据表对象获取表的名字
# sheet_name1 = sheet.title
# sheet_name2 = sheet2.title
# print(sheet_name1, sheet_name2)  # 结果：sheet1 sheet2

# 1.获取整个一行的单元格
max_column = sheet.max_column  # 获取最大列数
column = get_column_letter(max_column)  # 获取最大列数对应的字母列号
# 获取第一行所有单元格对象
row2 = sheet['A1':'%s1' % column]  # ((<Cell '表1'.A1>, <Cell '表1'.B1>, <Cell '表1'.C1>),)

for row_cells in row2:
    for cell in row_cells:
        print(cell.coordinate, cell.value)

# 获取整个列的单元格
max_row = sheet.max_row
columnB = sheet['B1':'B%d' % max_row]
# 获取B列对应的所有单元格对象
for column_cells in columnB:
    for cell in column_cells:
        print(cell.coordinate, cell.value)

# 获取矩形区域中的单元格对象
cell_tuples = sheet['A1': 'C3']
for cells in cell_tuples:
    for cell in cells:
        print(cell.coordinate, cell.value)
